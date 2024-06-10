from flask import Flask, request, render_template, send_file
import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
import io
from routes import CompetitorAnalysis

app = Flask(__name__)
app.register_blueprint(CompetitorAnalysis, url_prefix='')


@app.route('/')
def index():
    return render_template('index.html')


# def generate_swot_analysis(competitor_name, key_metrics):
#     # Placeholder function for SWOT analysis
#     return {
#         "market_positioning": "High",
#         "strengths": ["Strong brand", "Large market share"],
#         "weaknesses": ["High costs", "Dependence on few products"],
#         "opportunities": ["Emerging markets", "New technologies"],
#         "threats": ["Competition", "Regulatory changes"]
#     }

# def create_pdf_report(analysis):
#     pdf = FPDF()
#     pdf.add_page()
#     pdf.set_font("Arial", size=12)
#     for key, value in analysis.items():
#         pdf.cell(200, 10, txt=f"{key}: {value}", ln=True)
#     output = io.BytesIO()
#     pdf.output(output)
#     output.seek(0)
#     return output

# def create_excel_report(analysis):
#     workbook = Workbook()
#     sheet = workbook.active
#     row = 1
#     for key, value in analysis.items():
#         sheet.cell(row=row, column=1, value=key)
#         sheet.cell(row=row, column=2, value=str(value))
#         row += 1
#     output = io.BytesIO()
#     workbook.save(output)
#     output.seek(0)
#     return output

# @app.route('/', methods=['GET', 'POST'])
# def index():
#     if request.method == 'POST':
#         competitor_name = request.form['competitor_name']
#         key_metrics = request.form['key_metrics']
        
#         analysis = generate_swot_analysis(competitor_name, key_metrics)
        
#         if 'download_pdf' in request.form:
#             pdf = create_pdf_report(analysis)
#             return send_file(pdf, as_attachment=True, download_name='analysis_report.pdf', mimetype='application/pdf')
        
#         if 'download_excel' in request.form:
#             excel = create_excel_report(analysis)
#             return send_file(excel, as_attachment=True, download_name='analysis_report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
#         return render_template('index.html', analysis=analysis)
    
#     return render_template('index.html', analysis=None)

if __name__ == '__main__':
    app.run(debug=True)
